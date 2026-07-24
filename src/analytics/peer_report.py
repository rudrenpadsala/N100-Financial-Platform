# ==========================================================
# Peer Comparison Excel Report
# Day 20
# ==========================================================

import sqlite3
from pathlib import Path

import pandas as pd

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font


# ----------------------------------------------------------
# Paths
# ----------------------------------------------------------

DB_PATH = "db/nifty100.db"

OUTPUT_FILE = "output/peer_comparison.xlsx"

Path("output").mkdir(exist_ok=True)

# ==========================================================
# Load Database Tables
# ==========================================================

def load_tables():

    conn = sqlite3.connect(DB_PATH)

    companies = pd.read_sql(
        """
        SELECT
            id,
            company_name
        FROM companies
        """,
        conn
    )

    peer_groups = pd.read_sql(
        """
        SELECT
            company_id,
            peer_group_name,
            is_benchmark
        FROM peer_groups
        """,
        conn
    )

    financial = pd.read_sql(
        """
        SELECT *
        FROM financial_ratios
        """,
        conn
    )

    percentiles = pd.read_sql(
        """
        SELECT *
        FROM peer_percentiles
        """,
        conn
    )

    conn.close()

    print("\n" + "=" * 60)
    print("DATABASE LOADED")
    print("=" * 60)

    print("Companies       :", len(companies))
    print("Peer Groups     :", len(peer_groups))
    print("Financial Rows  :", len(financial))
    print("Percentile Rows :", len(percentiles))

    return (
        companies,
        peer_groups,
        financial,
        percentiles
    )

# ==========================================================
# Preview Tables
# ==========================================================

def preview_tables():

    (
        companies,
        peer_groups,
        financial,
        percentiles
    ) = load_tables()

    print("\nCompanies")
    print(companies.head())

    print("\nPeer Groups")
    print(peer_groups.head())

    print("\nFinancial Ratios")
    print(financial.head())

    print("\nPeer Percentiles")
    print(percentiles.head())

    return (
        companies,
        peer_groups,
        financial,
        percentiles
    )

# ==========================================================
# Latest Financial Data
# ==========================================================

def latest_financial_data(financial):

    """
    Keep only the latest year for every company.
    """

    latest = (
        financial
        .sort_values("year")
        .groupby("company_id")
        .tail(1)
        .reset_index(drop=True)
    )

    print("\n" + "=" * 60)
    print("LATEST FINANCIAL DATA")
    print("=" * 60)

    print("Companies :", len(latest))

    return latest
# ==========================================================
# Pivot Percentile Table
# ==========================================================

def pivot_percentiles(percentiles):

    """
    Convert metric rows into columns.
    """

    pivot = (
        percentiles
        .pivot_table(
            index=[
                "company_id",
                "peer_group_name",
                "year"
            ],
            columns="metric",
            values="percentile_rank"
        )
        .reset_index()
    )

    pivot.columns.name = None

    print("\n" + "=" * 60)
    print("PERCENTILE TABLE")
    print("=" * 60)

    print("Rows :", len(pivot))

    return pivot

# ==========================================================
# Prepare Final Report
# ==========================================================

def prepare_report():

    (
        companies,
        peer_groups,
        financial,
        percentiles
    ) = load_tables()

    financial = latest_financial_data(financial)

    percentile_table = pivot_percentiles(percentiles)

    report = (
        financial
        .merge(
            companies,
            left_on="company_id",
            right_on="id",
            how="left"
        )
        .merge(
            peer_groups,
            on="company_id",
            how="left"
        )
        .merge(
            percentile_table,
            on=[
                "company_id",
                "peer_group_name",
                "year"
            ],
            how="left"
        )
    )

    print("\n" + "=" * 60)
    print("FINAL REPORT")
    print("=" * 60)

    print("Rows :", len(report))
    print("Columns :", len(report.columns))

    return report

# ==========================================================
# Preview Report
# ==========================================================

def preview_report():

    report = prepare_report()

    print()

    print(report.head())

    print()

    print(report.columns.tolist())

    return report

# ==========================================================
# Export Excel
# ==========================================================

# ==========================================================
# Export Peer Comparison Excel
# ==========================================================

def export_peer_report(report):

    """
    Generate peer_comparison.xlsx

    One worksheet per peer group.
    """

    

    with pd.ExcelWriter(
        OUTPUT_FILE,
        engine="openpyxl"
    ) as writer:

        peer_groups = sorted(
            report["peer_group_name"]
            .dropna()
            .unique()
        )

        for peer in peer_groups:

            temp = (
                report[
                    report["peer_group_name"] == peer
                ]
                .copy()
            )

            # ------------------------------------------
            # Sort Companies
            # ------------------------------------------

            temp = temp.sort_values(
                "company_name"
            )

            # ------------------------------------------
            # Median Summary Row
            # ------------------------------------------

            numeric_cols = temp.select_dtypes(
                include="number"
            ).columns

            median_row = {}

            for col in temp.columns:

                if col in numeric_cols:

                    median_row[col] = round(
                        temp[col].median(),
                        2
                    )

                else:

                    median_row[col] = ""

            median_row["company_name"] = "PEER GROUP MEDIAN"

            temp = pd.concat(
                [
                    temp,
                    pd.DataFrame([median_row])
                ],
                ignore_index=True
            )

            # ------------------------------------------
            # Excel Sheet
            # ------------------------------------------

            sheet_name = peer[:31]

            temp.to_excel(
                writer,
                sheet_name=sheet_name,
                index=False
            )

    print("\n" + "=" * 60)
    print("Peer Comparison Excel Generated")
    print("=" * 60)
    print(f"Saved : {OUTPUT_FILE}")

    return OUTPUT_FILE

# ==========================================================
# Format Excel Report
# ==========================================================

# ==========================================================
# Format Peer Comparison Excel
# ==========================================================

def format_peer_report():

    """
    Apply formatting to peer_comparison.xlsx
    """

    wb = load_workbook(OUTPUT_FILE)

    # --------------------------------------------------
    # Cell Styles
    # --------------------------------------------------

    green_fill = PatternFill(
        fill_type="solid",
        start_color="C6EFCE"
    )

    yellow_fill = PatternFill(
        fill_type="solid",
        start_color="FFF2CC"
    )

    red_fill = PatternFill(
        fill_type="solid",
        start_color="F4CCCC"
    )

    benchmark_fill = PatternFill(
        fill_type="solid",
        start_color="FFD966"
    )

    median_fill = PatternFill(
        fill_type="solid",
        start_color="D9EAD3"
    )

    bold_font = Font(
        bold=True
    )

    # --------------------------------------------------
    # Every Worksheet
    # --------------------------------------------------

    for ws in wb.worksheets:

        headers = {}

        for col in range(1, ws.max_column + 1):

            headers[
                ws.cell(1, col).value
            ] = col

        # ------------------------------------------
        # Bold Header
        # ------------------------------------------

        for cell in ws[1]:

            cell.font = bold_font

        company_col = headers.get("company_name")

        benchmark_col = headers.get("is_benchmark")

        # ------------------------------------------
        # Find Percentile Columns
        # ------------------------------------------

        percentile_columns = []

        percentile_names = [

            "ROE",
            "ROCE",
            "Net Profit Margin",
            "Debt to Equity",
            "Free Cash Flow",
            "PAT CAGR",
            "Revenue CAGR",
            "EPS CAGR",
            "Interest Coverage",
            "Asset Turnover"

        ]

        for name in percentile_names:

            if name in headers:

                percentile_columns.append(
                    headers[name]
                )

        # ------------------------------------------
        # Row Formatting
        # ------------------------------------------

        for row in range(2, ws.max_row + 1):

            # ----------------------------
            # Benchmark Company
            # ----------------------------

            if benchmark_col:

                value = ws.cell(
                    row,
                    benchmark_col
                ).value

                if value == 1:

                    for col in range(
                        1,
                        ws.max_column + 1
                    ):

                        cell = ws.cell(
                            row,
                            col
                        )

                        cell.fill = benchmark_fill
                        cell.font = bold_font

            # ----------------------------
            # Median Row
            # ----------------------------

            if company_col:

                text = ws.cell(
                    row,
                    company_col
                ).value

                if text == "PEER GROUP MEDIAN":

                    for col in range(
                        1,
                        ws.max_column + 1
                    ):

                        cell = ws.cell(
                            row,
                            col
                        )

                        cell.fill = median_fill
                        cell.font = bold_font

            # ----------------------------
            # Percentile Colors
            # ----------------------------

            for col in percentile_columns:

                value = ws.cell(
                    row,
                    col
                ).value

                if isinstance(
                    value,
                    (int, float)
                ):

                    if value >= 75:

                        ws.cell(
                            row,
                            col
                        ).fill = green_fill

                    elif value <= 25:

                        ws.cell(
                            row,
                            col
                        ).fill = red_fill

                    else:

                        ws.cell(
                            row,
                            col
                        ).fill = yellow_fill

        # ------------------------------------------
        # Auto Width
        # ------------------------------------------

        for column in ws.columns:

            length = 0

            column_letter = column[0].column_letter

            for cell in column:

                try:

                    if len(str(cell.value)) > length:

                        length = len(str(cell.value))

                except:

                    pass

            ws.column_dimensions[
                column_letter
            ].width = min(
                length + 3,
                30
            )

    wb.save(OUTPUT_FILE)

    print("\n" + "=" * 60)
    print("Excel Formatting Completed")
    print("=" * 60)
    print(f"Saved : {OUTPUT_FILE}")


# ==========================================================
# Main
# ==========================================================

# ==========================================================
# Main
# ==========================================================

def main():

    print("=" * 60)
    print("Peer Comparison Report")
    print("=" * 60)

    # ------------------------------------------
    # Preview Database
    # ------------------------------------------

    preview_tables()

    # ------------------------------------------
    # Prepare Final Report (Only Once)
    # ------------------------------------------

    report = prepare_report()

    # ------------------------------------------
    # Export Excel Report
    # ------------------------------------------

    export_peer_report(report)

    # ------------------------------------------
    # Apply Excel Formatting
    # ------------------------------------------

    format_peer_report()

    # ------------------------------------------
    # Final Summary
    # ------------------------------------------

    print("\n" + "=" * 60)
    print("DAY 20 SUMMARY")
    print("=" * 60)

    print(f"Companies             : {report.shape[0]}")
    print(f"Peer Groups           : {report['peer_group_name'].dropna().nunique()}")
    print(f"Benchmark Companies   : {report['is_benchmark'].fillna(0).sum():.0f}")
    print(f"Output File           : {OUTPUT_FILE}")

    print("\n" + "=" * 60)
    print("Day 20 Completed Successfully")
    print("=" * 60)


# ==========================================================
# Run
# ==========================================================

if __name__ == "__main__":
    main()