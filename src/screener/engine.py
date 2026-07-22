# ==========================================================
# Financial Screener Engine
# N100 Financial Intelligence Platform
# ==========================================================

import sqlite3
from pathlib import Path

import pandas as pd
import yaml

from openpyxl import load_workbook
from openpyxl.styles import PatternFill


# ==========================================================
# Paths
# ==========================================================

DB_PATH = "db/nifty100.db"

CONFIG_PATH = "src/config/screener_config.yaml"

OUTPUT_FILE = "output/screener_output.xlsx"


# ==========================================================
# Load Data
# ==========================================================

def load_data():
    """
    Load all required tables from SQLite database.
    """

    conn = sqlite3.connect(DB_PATH)

    # ------------------------------------------------------
    # Financial Ratios
    # ------------------------------------------------------

    financial_ratios = pd.read_sql(
        """
        SELECT *
        FROM financial_ratios
        """,
        conn
    )

    # ------------------------------------------------------
    # Company Master
    # ------------------------------------------------------

    companies = pd.read_sql(
        """
        SELECT *
        FROM companies
        """,
        conn
    )

    # ------------------------------------------------------
    # Sector Information
    # ------------------------------------------------------

    sectors = pd.read_sql(
        """
        SELECT
            company_id,
            broad_sector,
            sub_sector,
            market_cap_category
        FROM sectors
        """,
        conn
    )

    # ------------------------------------------------------
    # Market Valuation
    # ------------------------------------------------------

    market_cap = pd.read_sql(
        """
        SELECT
            company_id,
            year,
            market_cap_crore,
            enterprise_value_crore,
            pe_ratio,
            pb_ratio,
            dividend_yield_pct
        FROM market_cap
        """,
        conn
    )

    # ------------------------------------------------------
    # Analysis Table
    # ------------------------------------------------------

    analysis = pd.read_sql(
        """
        SELECT
            company_id,
            compounded_sales_growth,
            compounded_profit_growth,
            stock_price_cagr,
            roe
        FROM analysis
        """,
        conn
    )

    conn.close()

    return (
        financial_ratios,
        companies,
        sectors,
        market_cap,
        analysis
    )


# ==========================================================
# Load Screener Configuration
# ==========================================================

def load_config():
    """
    Load screener presets from YAML configuration.
    """

    with open(
        CONFIG_PATH,
        "r",
        encoding="utf-8"
    ) as file:

        config = yaml.safe_load(file)

    return config

# ==========================================================
# Apply Screener Filters
# ==========================================================

def apply_filters(df, filters):
    """
    Apply screener filters from YAML configuration.
    """

    filtered = df.copy()

    # ======================================================
    # Return on Equity
    # ======================================================

    if "roe_min" in filters:

        filtered = filtered[
            filtered["return_on_equity_pct"] >= filters["roe_min"]
        ]

    # ======================================================
    # Debt to Equity
    # Financial sector is ignored
    # ======================================================

    if "de_max" in filters:

        if "broad_sector" in filtered.columns:

            financials = filtered[
                filtered["broad_sector"] == "Financials"
            ]

            others = filtered[
                filtered["broad_sector"] != "Financials"
            ]

            others = others[
                (
                    others["debt_to_equity"].isna()
                )
                |
                (
                    others["debt_to_equity"]
                    <= filters["de_max"]
                )
            ]

            filtered = pd.concat(
                [financials, others],
                ignore_index=True
            )

        else:

            filtered = filtered[
                (
                    filtered["debt_to_equity"].isna()
                )
                |
                (
                    filtered["debt_to_equity"]
                    <= filters["de_max"]
                )
            ]

    # ======================================================
    # Free Cash Flow
    # ======================================================

    if "fcf_min" in filters:

        filtered = filtered[
            filtered["free_cash_flow_cr"]
            >= filters["fcf_min"]
        ]

    # ======================================================
    # Revenue CAGR (5 Year)
    # ======================================================

    if "revenue_cagr_min" in filters:

        filtered = filtered[
            filtered["revenue_cagr_5yr"]
            >= filters["revenue_cagr_min"]
        ]

    # ======================================================
    # Revenue CAGR 3Y
    # Dataset only has 5Y CAGR
    # ======================================================

    if "revenue_cagr_3yr_min" in filters:

        filtered = filtered[
            filtered["revenue_cagr_5yr"]
            >= filters["revenue_cagr_3yr_min"]
        ]

    # ======================================================
    # PAT CAGR
    # ======================================================

    if "pat_cagr_min" in filters:

        filtered = filtered[
            filtered["pat_cagr_5yr"]
            >= filters["pat_cagr_min"]
        ]

    # ======================================================
    # Operating Profit Margin
    # ======================================================

    if "opm_min" in filters:

        filtered = filtered[
            filtered["operating_profit_margin_pct"]
            >= filters["opm_min"]
        ]

    # ======================================================
    # Interest Coverage
    # ======================================================

    if "icr_min" in filters:

        filtered = filtered[
            (
                filtered["interest_coverage"].isna()
            )
            |
            (
                filtered["interest_coverage"]
                >= filters["icr_min"]
            )
        ]

    # ======================================================
    # PE Ratio
    # ======================================================

    if (
        "pe_max" in filters
        and "pe_ratio" in filtered.columns
    ):

        filtered = filtered[
            (
                filtered["pe_ratio"].isna()
            )
            |
            (
                filtered["pe_ratio"]
                <= filters["pe_max"]
            )
        ]

    # ======================================================
    # PB Ratio
    # ======================================================

    if (
        "pb_max" in filters
        and "pb_ratio" in filtered.columns
    ):

        filtered = filtered[
            (
                filtered["pb_ratio"].isna()
            )
            |
            (
                filtered["pb_ratio"]
                <= filters["pb_max"]
            )
        ]

    # ======================================================
    # Dividend Yield
    # ======================================================

    if (
        "dividend_yield_min" in filters
        and "dividend_yield_pct" in filtered.columns
    ):

        filtered = filtered[
            (
                filtered["dividend_yield_pct"].isna()
            )
            |
            (
                filtered["dividend_yield_pct"]
                >= filters["dividend_yield_min"]
            )
        ]

    # ======================================================
    # Dividend Payout Ratio
    # ======================================================

    if "dividend_payout_max" in filters:

        filtered = filtered[
            filtered["dividend_payout_ratio_pct"]
            <= filters["dividend_payout_max"]
        ]

    # ======================================================
    # Market Cap
    # ======================================================

    if (
        "market_cap_min" in filters
        and "market_cap_crore" in filtered.columns
    ):

        filtered = filtered[
            filtered["market_cap_crore"]
            >= filters["market_cap_min"]
        ]

    # ======================================================
    # EPS CAGR
    # ======================================================

    if (
        "eps_cagr_min" in filters
        and "eps_cagr_5yr" in filtered.columns
    ):

        filtered = filtered[
            filtered["eps_cagr_5yr"]
            >= filters["eps_cagr_min"]
        ]

    # ======================================================
    # Asset Turnover
    # ======================================================

    if (
        "asset_turnover_min" in filters
        and "asset_turnover" in filtered.columns
    ):

        filtered = filtered[
            filtered["asset_turnover"]
            >= filters["asset_turnover_min"]
        ]

    # ======================================================
    # NOTE:
    # sales_min filter removed because your database
    # has no "sales" column.
    # ======================================================

    return filtered

# ==========================================================
# Winsorize + Scale
# ==========================================================

def winsorize_and_scale(series):
    """
    Winsorize using P10/P90 and scale to 0-100.
    """

    s = pd.to_numeric(series, errors="coerce").fillna(0)

    p10 = s.quantile(0.10)
    p90 = s.quantile(0.90)

    s = s.clip(lower=p10, upper=p90)

    if p90 == p10:
        return pd.Series(50, index=s.index)

    return ((s - p10) / (p90 - p10)) * 100


# ==========================================================
# Composite Quality Score
# ==========================================================

def calculate_composite_score(df):
    """
    Calculate Composite Quality Score.

    Weightage
    ----------
    Profitability : 35%
    Cash Quality  : 30%
    Growth        : 20%
    Leverage      : 15%
    """

    df = df.copy()

    score = pd.Series(0.0, index=df.index)

    # =====================================================
    # PROFITABILITY (35%)
    # =====================================================

    score += winsorize_and_scale(
        df["return_on_equity_pct"]
    ) * 0.15

    score += winsorize_and_scale(
        df["return_on_capital_employed_pct"]
    ) * 0.10

    score += winsorize_and_scale(
        df["net_profit_margin_pct"]
    ) * 0.10

    # =====================================================
    # CASH QUALITY (30%)
    # =====================================================

    if "fcf_cagr_5yr" in df.columns:

        score += winsorize_and_scale(
            df["fcf_cagr_5yr"]
        ) * 0.15

    if "cfo_quality_score" in df.columns:

        score += winsorize_and_scale(
            df["cfo_quality_score"]
        ) * 0.10

    if "free_cash_flow_cr" in df.columns:

        score += (
            (df["free_cash_flow_cr"] > 0)
            .astype(float)
            * 5
        )

    # =====================================================
    # GROWTH (20%)
    # =====================================================

    score += winsorize_and_scale(
        df["revenue_cagr_5yr"]
    ) * 0.10

    score += winsorize_and_scale(
        df["pat_cagr_5yr"]
    ) * 0.10

    # =====================================================
    # LEVERAGE (15%)
    # =====================================================

    score += winsorize_and_scale(
        -df["debt_to_equity"].fillna(0)
    ) * 0.10

    score += winsorize_and_scale(
        df["interest_coverage"]
    ) * 0.05

    # =====================================================
    # Normalize Final Score
    # =====================================================

    score = score.fillna(0)

    if score.max() != score.min():

        score = (
            (score - score.min())
            /
            (score.max() - score.min())
        ) * 100

    else:

        score = pd.Series(
            50,
            index=score.index
        )

    df["composite_quality_score"] = score.round(2)

    return df

# ==========================================================
# Composite Quality Score
# ==========================================================

def calculate_composite_score(df):
    """
    Calculate Composite Quality Score (0-100)

    Profitability : 35%
    Cash Quality  : 30%
    Growth        : 20%
    Leverage      : 15%
    """

    df = df.copy()

    score = pd.Series(0.0, index=df.index)

    # =====================================================
    # PROFITABILITY (35%)
    # =====================================================

    if "return_on_equity_pct" in df.columns:

        score += (
            winsorize_and_scale(
                df["return_on_equity_pct"]
            ) * 0.15
        )

    # companies table contains ROCE

    if "roce_percentage" in df.columns:

        score += (
            winsorize_and_scale(
                df["roce_percentage"]
            ) * 0.10
        )

    if "net_profit_margin_pct" in df.columns:

        score += (
            winsorize_and_scale(
                df["net_profit_margin_pct"]
            ) * 0.10
        )

    # =====================================================
    # CASH QUALITY (30%)
    # =====================================================

    if "free_cash_flow_cr" in df.columns:

        score += (
            winsorize_and_scale(
                df["free_cash_flow_cr"]
            ) * 0.20
        )

    if "cash_from_operations_cr" in df.columns:

        score += (
            winsorize_and_scale(
                df["cash_from_operations_cr"]
            ) * 0.10
        )

    # =====================================================
    # GROWTH (20%)
    # =====================================================

    if "revenue_cagr_5yr" in df.columns:

        score += (
            winsorize_and_scale(
                df["revenue_cagr_5yr"]
            ) * 0.10
        )

    if "pat_cagr_5yr" in df.columns:

        score += (
            winsorize_and_scale(
                df["pat_cagr_5yr"]
            ) * 0.10
        )

    # =====================================================
    # LEVERAGE (15%)
    # =====================================================

    if "debt_to_equity" in df.columns:

        score += (
            winsorize_and_scale(
                -df["debt_to_equity"].fillna(0)
            ) * 0.10
        )

    if "interest_coverage" in df.columns:

        score += (
            winsorize_and_scale(
                df["interest_coverage"]
            ) * 0.05
        )

    # =====================================================
    # Normalize Composite Score
    # =====================================================

    score = score.fillna(0)

    if score.max() > score.min():

        score = (
            (score - score.min())
            /
            (score.max() - score.min())
        ) * 100

    else:

        score = pd.Series(
            50,
            index=score.index
        )

    df["composite_quality_score"] = score.round(2)

    # =====================================================
    # Sector Relative Score
    # =====================================================

    if "broad_sector" in df.columns:

        df["sector_relative_score"] = (
            df.groupby("broad_sector")[
                "composite_quality_score"
            ]
            .transform(
                lambda x: (
                    (
                        x - x.min()
                    )
                    /
                    (
                        x.max() - x.min()
                    )
                    * 100
                )
                if x.max() != x.min()
                else 50
            )
        ).round(2)

    else:

        df["sector_relative_score"] = (
            df["composite_quality_score"]
        )

    return df

# ==========================================================
# Run Screener
# ==========================================================

def run_screener(filters):
    """
    Run screener for a single preset.
    """

    # ------------------------------------------------------
    # Load Data
    # ------------------------------------------------------

    (
        financial_ratios,
        companies,
        sectors,
        market_cap,
        analysis
    ) = load_data()

    # ------------------------------------------------------
    # Merge Financial Ratios + Companies
    # ------------------------------------------------------

    df = financial_ratios.merge(
        companies,
        left_on="company_id",
        right_on="id",
        how="left",
        suffixes=("", "_company")
    )

    # ------------------------------------------------------
    # Merge Sector Table
    # ------------------------------------------------------

    df = df.merge(
        sectors,
        on="company_id",
        how="left"
    )

    # ------------------------------------------------------
    # Merge Market Cap Table
    # ------------------------------------------------------

    df = df.merge(
        market_cap[
            [
                "company_id",
                "market_cap_crore",
                "pe_ratio",
                "pb_ratio",
                "dividend_yield_pct"
            ]
        ],
        on="company_id",
        how="left"
    )

    # ------------------------------------------------------
    # Merge Analysis Table
    # ------------------------------------------------------

    df = df.merge(
        analysis[
            [
                "company_id",
                "compounded_sales_growth",
                "compounded_profit_growth"
            ]
        ],
        on="company_id",
        how="left"
    )

    # ------------------------------------------------------
    # Rename columns
    # ------------------------------------------------------

    df.rename(
        columns={

            "market_cap_crore": "market_cap",

            "pe_ratio": "price_to_earnings",

            "pb_ratio": "price_to_book",

            "dividend_yield_pct": "dividend_yield",

            "compounded_sales_growth": "revenue_cagr_3yr",

            "compounded_profit_growth": "pat_cagr_3yr"

        },
        inplace=True
    )

    # ------------------------------------------------------
    # Keep Latest Year
    # ------------------------------------------------------

    if "year" in df.columns:

        df = (
            df.sort_values("year")
              .groupby("company_id")
              .tail(1)
              .reset_index(drop=True)
        )

    # ------------------------------------------------------
    # Composite Score
    # ------------------------------------------------------

    df = calculate_composite_score(df)

    # ------------------------------------------------------
    # Apply Filters
    # ------------------------------------------------------

    df = apply_filters(
        df,
        filters
    )

    # ------------------------------------------------------
    # Sort Results
    # ------------------------------------------------------

    if "composite_quality_score" in df.columns:

        df = df.sort_values(
            by="composite_quality_score",
            ascending=False
        )

    df = df.reset_index(drop=True)

    return df

# ==========================================================
# Export Presets
# ==========================================================

def export_presets(results):
    """
    Export every screener preset to Excel.
    One worksheet per preset.
    """

    output_dir = Path("output")
    output_dir.mkdir(exist_ok=True)

    output_file = output_dir / "screener_output.xlsx"

    with pd.ExcelWriter(
        output_file,
        engine="openpyxl"
    ) as writer:

        for preset_name, df in results.items():

            export_df = df.copy()

            # --------------------------------------------------
            # Columns to Export
            # --------------------------------------------------

            export_columns = [

                "company_id",
                "company_name",

                "broad_sector",
                "sub_sector",

                "year",

                "return_on_equity_pct",
                "return_on_capital_employed_pct",
                "net_profit_margin_pct",

                "free_cash_flow_cr",

                "revenue_cagr_5yr",
                "pat_cagr_5yr",

                "revenue_cagr_3yr",
                "pat_cagr_3yr",

                "debt_to_equity",
                "interest_coverage",

                "price_to_earnings",
                "price_to_book",

                "market_cap",
                "dividend_yield",
                "dividend_payout_ratio_pct",

                "composite_quality_score",
                "sector_relative_score"

            ]

            # Keep only existing columns

            export_columns = [
                col
                for col in export_columns
                if col in export_df.columns
            ]

            export_df = export_df[export_columns]

            # --------------------------------------------------
            # Sort by Composite Score
            # --------------------------------------------------

            if "composite_quality_score" in export_df.columns:

                export_df = export_df.sort_values(
                    by="composite_quality_score",
                    ascending=False
                )

            # --------------------------------------------------
            # Export Sheet
            # --------------------------------------------------

            export_df.to_excel(
                writer,
                sheet_name=preset_name[:31],
                index=False
            )

    print(f"\nExcel saved : {output_file}")

# ==========================================================
# Excel Formatting
# ==========================================================

def format_excel():
    """
    Apply conditional formatting to the exported workbook.
    """

    workbook = load_workbook(OUTPUT_FILE)

    green = PatternFill(
        fill_type="solid",
        start_color="C6EFCE",
        end_color="C6EFCE"
    )

    red = PatternFill(
        fill_type="solid",
        start_color="FFC7CE",
        end_color="FFC7CE"
    )

    threshold_map = {

        "return_on_equity_pct":15,
        "return_on_capital_employed_pct":15,
        "net_profit_margin_pct":10,

        "revenue_cagr_5yr":10,
        "pat_cagr_5yr":10,

        "interest_coverage":3,

        "debt_to_equity":1,

        "price_to_earnings":30,
        "price_to_book":5,

        "dividend_yield":1,
        "dividend_payout_ratio_pct":80,

        "composite_quality_score":70,
        "sector_relative_score":70
    }

    lower_is_better = {

        "debt_to_equity",
        "price_to_earnings",
        "price_to_book",
        "dividend_payout_ratio_pct"

    }

    for sheet in workbook.sheetnames:

        ws = workbook[sheet]

        headers = {
            cell.value: cell.column
            for cell in ws[1]
        }

        for metric, threshold in threshold_map.items():

            if metric not in headers:
                continue

            col = headers[metric]

            for row in range(2, ws.max_row + 1):

                cell = ws.cell(row=row, column=col)

                value = cell.value

                if value is None:
                    continue

                # -----------------------------
                # Convert text to number safely
                # -----------------------------

                try:
                    value = float(value)
                except (ValueError, TypeError):
                    continue

                if metric in lower_is_better:

                    if value <= threshold:
                        cell.fill = green
                    else:
                        cell.fill = red

                else:

                    if value >= threshold:
                        cell.fill = green
                    else:
                        cell.fill = red

    workbook.save(OUTPUT_FILE)

    print("Excel formatting applied.")

# ==========================================================
# Run All Screener Presets
# ==========================================================

def run_all_presets():
    """
    Run all screener presets from YAML configuration.
    """

    config = load_config()

    results = {}

    print("\n" + "=" * 60)
    print("Running All Screener Presets")
    print("=" * 60)

    # ------------------------------------------------------
    # Run Every Preset
    # ------------------------------------------------------

    for preset_name, filters in config.items():

        print("\n" + "-" * 60)
        print(f"Preset : {preset_name}")
        print("-" * 60)

        try:

            df = run_screener(filters)

            print(f"Companies Found : {len(df)}")

            if not df.empty:

                print(
                    f"Highest Composite Score : "
                    f"{df['composite_quality_score'].max():.2f}"
                )

                if "sector_relative_score" in df.columns:

                    print(
                        f"Highest Sector Score    : "
                        f"{df['sector_relative_score'].max():.2f}"
                    )

            results[preset_name] = df

        except Exception as e:

            print(f"Error running {preset_name}")

            print(e)

            results[preset_name] = pd.DataFrame()

    # ------------------------------------------------------
    # Export Excel
    # ------------------------------------------------------

    export_presets(results)

    # ------------------------------------------------------
    # Apply Formatting
    # ------------------------------------------------------

    format_excel()

    print("\n" + "=" * 60)
    print("Excel Export Completed")
    print("=" * 60)

    return results
# ==========================================================
# Main
# ==========================================================

def main():
    """
    Main entry point for the Financial Screener.
    """

    print("=" * 60)
    print("Financial Screener")
    print("=" * 60)

    # ------------------------------------------------------
    # Run All Presets
    # ------------------------------------------------------

    all_results = run_all_presets()

    # ------------------------------------------------------
    # Display Results
    # ------------------------------------------------------

    for preset_name, df in all_results.items():

        print("\n" + "=" * 60)
        print(f"PRESET : {preset_name}")
        print("=" * 60)

        print(f"Companies Found : {len(df)}")

        if df.empty:

            print("No companies found.")
            continue

        display_columns = [

            "company_id",
            "company_name",

            "broad_sector",
            "sub_sector",

            "year",

            "return_on_equity_pct",
            "roce_percentage",
            "net_profit_margin_pct",

            "free_cash_flow_cr",
            "cash_from_operations_cr",

            "revenue_cagr_5yr",
            "pat_cagr_5yr",

            "price_to_earnings",
            "price_to_book",

            "debt_to_equity",
            "interest_coverage",

            "market_cap",

            "composite_quality_score",
            "sector_relative_score"

        ]

        display_columns = [

            col
            for col in display_columns
            if col in df.columns

        ]

        print("\nTop 10 Companies\n")

        print(

            df[display_columns]
            .head(10)
            .to_string(index=False)

        )

    # ------------------------------------------------------
    # Completed
    # ------------------------------------------------------

    print("\n" + "=" * 60)
    print("Financial Screener Completed Successfully")
    print("=" * 60)


# ==========================================================
# Entry Point
# ==========================================================

if __name__ == "__main__":
    main()
