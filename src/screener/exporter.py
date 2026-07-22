import os
import pandas as pd

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

OUTPUT_FILE = "output/screener_output.xlsx"

def export_presets(results):
    """
    Export all screener presets to Excel.
    """

    os.makedirs("output", exist_ok=True)

    with pd.ExcelWriter(
        OUTPUT_FILE,
        engine="openpyxl"
    ) as writer:

        for preset, df in results.items():

            df.to_excel(
                writer,
                sheet_name=preset[:31],
                index=False
            )

    color_excel()

    print(f"\nExcel saved : {OUTPUT_FILE}")

def color_excel():
    """
    Apply simple green/red formatting.
    """

    workbook = load_workbook(OUTPUT_FILE)

    green = PatternFill(
        fill_type="solid",
        start_color="C6EFCE",
        end_color="C6EFCE"
    )

    red = PatternFill(
        fill_type="solid",
        start_color="FFC7CE",
        end_color="FFC7CE"
    )

    for sheet in workbook.sheetnames:

        ws = workbook[sheet]

        headers = {}

        for cell in ws[1]:
            headers[cell.value] = cell.column

        for row in range(2, ws.max_row + 1):

            # ROE
            if "return_on_equity_pct" in headers:

                c = ws.cell(
                    row=row,
                    column=headers["return_on_equity_pct"]
                )

                if isinstance(c.value, (int, float)):

                    if c.value >= 15:
                        c.fill = green
                    else:
                        c.fill = red

            # Debt to Equity
            if "debt_to_equity" in headers:

                c = ws.cell(
                    row=row,
                    column=headers["debt_to_equity"]
                )

                if isinstance(c.value, (int, float)):

                    if c.value <= 1:
                        c.fill = green
                    else:
                        c.fill = red

            # Free Cash Flow
            if "free_cash_flow_cr" in headers:

                c = ws.cell(
                    row=row,
                    column=headers["free_cash_flow_cr"]
                )

                if isinstance(c.value, (int, float)):

                    if c.value > 0:
                        c.fill = green
                    else:
                        c.fill = red

            # Revenue CAGR
            if "revenue_cagr_5yr" in headers:

                c = ws.cell(
                    row=row,
                    column=headers["revenue_cagr_5yr"]
                )

                if isinstance(c.value, (int, float)):

                    if c.value >= 10:
                        c.fill = green
                    else:
                        c.fill = red

            # Composite Score
            if "composite_quality_score" in headers:

                c = ws.cell(
                    row=row,
                    column=headers["composite_quality_score"]
                )

                if isinstance(c.value, (int, float)):

                    if c.value >= 75:
                        c.fill = green
                    else:
                        c.fill = red

    workbook.save(OUTPUT_FILE)

    print("Excel formatting applied.")